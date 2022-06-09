from openpyxl import Workbook
wb = Workbook()

ws = wb.create_sheet() # 새로운 sheet 기본 이름으로 생성 (현재 활성화된 sheet 뒤에 새로운 sheet 생성)
ws.title = "MySheet" # sheet 이름 변경
ws.sheet_properties.tabColor = "ff66ff" # RGB 형태로 값을 넣어주면 탭 색상 변경

# Sheet, MySheet, YourSheet
ws1 = wb.create_sheet("YourSheet") # 주어진 이름으로 sheet 생성 (이것은 MySheet 뒤에 생성된다)

ws2 = wb.create_sheet("NewSheet", 2) # 2번째 인덱스에 sheet 생성

new_ws = wb["NewSheet"] # Dict(딕셔너리) 형태로 sheet에 접근 가능

print(wb.sheetnames) # 모든 sheet 이름 출력

# sheet 복사
new_ws["A1"] = "Test" # A1 셀의 값을 Test 로 설정
target = wb.copy_worksheet(new_ws)
target.title = "Copied Sheet" # new_ws을 카피해서 가장 마지막 시트에 Copied Sheet 라는 이름으로 추가

wb.save("sample.xlsx")